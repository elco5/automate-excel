from openpyxl import Workbook, load_workbook

wb = load_workbook('Grades.xlsx') # workbook must not be open
# get the worksheet from the workbook
ws = wb.active
print(ws)

print(ws['A1'].value)
